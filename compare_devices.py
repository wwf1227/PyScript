# -*- coding: utf-8 -*-
"""
对比 Excel 表格中两列设备机型的差异。

用法:
    python compare_devices.py 设备列表对比.xlsx
    python compare_devices.py 设备列表对比.xlsx -o 结果.xlsx --col-a 型号A --col-b 型号B
    python compare_devices.py 设备列表对比.xlsx --ignore-case   # 忽略大小写/空格

默认读取第一个 sheet 的前两列(含表头),输出一个新的 Excel,
包含 汇总 / 共有 / 仅在A列 / 仅在B列 四个 sheet。
"""
import argparse
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill

HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="4472C4")


def read_columns(path, sheet=None, col_a=None, col_b=None):
    """读取两列数据,返回 (列A标题, 列B标题, A值列表, B值列表)。

    col_a / col_b 可为列标题文字,也可为列序号(1 开始)。默认取前两列。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(values_only=True))]

    def resolve(spec, default_idx):
        if spec is None:
            return default_idx
        # 数字 -> 列序号(1 开始)
        if str(spec).isdigit():
            return int(spec) - 1
        # 否则按标题匹配
        if spec in header:
            return header.index(spec)
        raise ValueError("找不到列: %r,可用列: %s" % (spec, header))

    ia, ib = resolve(col_a, 0), resolve(col_b, 1)
    title_a = header[ia] if ia < len(header) else "列A"
    title_b = header[ib] if ib < len(header) else "列B"

    vals_a, vals_b = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if ia < len(row) and row[ia] not in (None, ""):
            vals_a.append(str(row[ia]).strip())
        if ib < len(row) and row[ib] not in (None, ""):
            vals_b.append(str(row[ib]).strip())
    return title_a, title_b, vals_a, vals_b


def compare(vals_a, vals_b, ignore_case=False):
    """对比两列,返回 dict:共有 / 仅A / 仅B / A列重复 / B列重复 / 写法不一致。"""
    norm = (lambda s: s.replace(" ", "").upper()) if ignore_case else (lambda s: s)

    map_a, map_b = {}, {}
    for v in vals_a:
        map_a.setdefault(norm(v), v)
    for v in vals_b:
        map_b.setdefault(norm(v), v)
    ka, kb = set(map_a), set(map_b)

    dup_a = sorted({x for x in vals_a if vals_a.count(x) > 1})
    dup_b = sorted({x for x in vals_b if vals_b.count(x) > 1})

    both_keys = sorted(ka & kb)

    # 以 A 列原始顺序逐个标注是否在 B 列(已测试)中,去重
    annotated, seen = [], set()
    for v in vals_a:
        k = norm(v)
        if k in seen:
            continue
        seen.add(k)
        annotated.append((v, "已测试" if k in kb else "未测试"))

    return {
        "both": [map_a[k] for k in both_keys],
        "only_a": [map_a[k] for k in sorted(ka - kb)],
        "only_b": [map_b[k] for k in sorted(kb - ka)],
        "dup_a": dup_a,
        "dup_b": dup_b,
        # A 列每个机型的测试状态:[(机型, '已测试'|'未测试'), ...]
        "annotated_a": annotated,
        # 忽略大小写/空格时,两列写法不同但归一化后相同的配对
        "mismatch": [(map_a[k], map_b[k]) for k in both_keys if map_a[k] != map_b[k]],
    }


def write_result(out_path, title_a, title_b, vals_a, vals_b, result, ignore_case):
    wb = openpyxl.Workbook()

    def header_cell(cell, text):
        cell.value = text
        cell.font = HDR_FONT
        cell.fill = HDR_FILL

    def one_col_sheet(name, items):
        s = wb.create_sheet(name)
        header_cell(s.cell(1, 1), "%s (%d)" % (name, len(items)))
        for i, v in enumerate(items, start=2):
            s.cell(i, 1, v)
        s.column_dimensions["A"].width = 24

    # 汇总
    sm = wb.active
    sm.title = "汇总"
    header_cell(sm["A1"], "类别")
    header_cell(sm["B1"], "数量")
    summary = [
        ("共有(两列都有)", len(result["both"])),
        ("仅在A列 (%s)" % title_a, len(result["only_a"])),
        ("仅在B列 (%s)" % title_b, len(result["only_b"])),
        ("A列未测试数量", len(result["only_a"])),
        ("%s 总数" % title_a, len(vals_a)),
        ("%s 总数" % title_b, len(vals_b)),
        ("%s 列内重复" % title_a, len(result["dup_a"])),
        ("%s 列内重复" % title_b, len(result["dup_b"])),
        ("忽略大小写/空格", "是" if ignore_case else "否"),
    ]
    for i, (k, v) in enumerate(summary, start=2):
        sm.cell(i, 1, k)
        sm.cell(i, 2, v)
    sm.column_dimensions["A"].width = 26
    sm.column_dimensions["B"].width = 12

    # A列测试标注:A列全量机型 + 是否在 B(已测试)列,未测试行标红
    s = wb.create_sheet("A列测试标注")
    header_cell(s.cell(1, 1), title_a)
    header_cell(s.cell(1, 2), "备注(是否测试)")
    untested_fill = PatternFill("solid", fgColor="FFC7CE")  # 浅红
    untested_font = Font(color="9C0006")
    for i, (v, status) in enumerate(result["annotated_a"], start=2):
        s.cell(i, 1, v)
        c = s.cell(i, 2, status)
        if status == "未测试":
            s.cell(i, 1).fill = untested_fill
            c.fill = untested_fill
            c.font = untested_font
    s.column_dimensions["A"].width = 22
    s.column_dimensions["B"].width = 16

    one_col_sheet("共有", result["both"])
    one_col_sheet("仅在A列", result["only_a"])
    one_col_sheet("仅在B列", result["only_b"])

    # 仅在忽略大小写时,且存在写法差异时输出
    if result["mismatch"]:
        s = wb.create_sheet("写法不一致")
        header_cell(s.cell(1, 1), "A列写法")
        header_cell(s.cell(1, 2), "B列写法")
        for i, (a, b) in enumerate(result["mismatch"], start=2):
            s.cell(i, 1, a)
            s.cell(i, 2, b)
        s.column_dimensions["A"].width = 22
        s.column_dimensions["B"].width = 22

    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(description="对比 Excel 中两列设备机型的差异")
    ap.add_argument("input", help="输入 Excel 文件路径")
    ap.add_argument("-o", "--output", help="输出文件路径(默认在输入文件名后加 _差异结果)")
    ap.add_argument("--sheet", help="sheet 名称(默认第一个)")
    ap.add_argument("--col-a", help="A列:列标题或列号(从1开始,默认第1列)")
    ap.add_argument("--col-b", help="B列:列标题或列号(从1开始,默认第2列)")
    ap.add_argument("--ignore-case", action="store_true", help="忽略大小写和空格再对比")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.exit("文件不存在: %s" % args.input)

    title_a, title_b, vals_a, vals_b = read_columns(
        args.input, args.sheet, args.col_a, args.col_b
    )
    result = compare(vals_a, vals_b, args.ignore_case)

    out = args.output
    if not out:
        base, ext = os.path.splitext(args.input)
        out = base + "_差异结果" + (ext or ".xlsx")

    write_result(out, title_a, title_b, vals_a, vals_b, result, args.ignore_case)

    print("对比列: A=%r  B=%r" % (title_a, title_b))
    print("共有 %d | 仅在A列 %d | 仅在B列 %d" % (
        len(result["both"]), len(result["only_a"]), len(result["only_b"])))
    print("A列共 %d 个机型,其中未测试 %d 个" % (
        len(result["annotated_a"]), len(result["only_a"])))
    if result["dup_a"] or result["dup_b"]:
        print("列内重复 -> A: %s | B: %s" % (result["dup_a"], result["dup_b"]))
    if result["mismatch"]:
        print("写法不一致(大小写/空格): %d 对" % len(result["mismatch"]))
    print("已生成: %s" % out)


if __name__ == "__main__":
    main()
