#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
@Time : 2026/06/29
@Author : wwf
Description: 从 xlsx 读取多列「秒.帧」文本，逐列转换为毫秒，结果写入新 sheet。
            原始数据 sheet 保持不动。

注意：单元格为文本格式，小数点后是帧数（不是小数）。
     例如 "1.10" 表示 1 秒 10 帧，必须按字符串拆分，
     绝不能当成浮点数读，否则 "1.10" 会被错读成 1.1（10 帧 -> 1 帧）。
"""

import openpyxl

# ===== 配置 =====
INPUT_FILE = "fps2time.xlsx"   # TODO: 改成实际的 xlsx 文件名
SOURCE_SHEET = None          # None=第一个 sheet；或填具体 sheet 名
RESULT_SHEET = "结果"        # 毫秒结果写入的 sheet 名（已存在则覆盖重建）
RESULT_SHEET_SEC = "结果(秒)"  # 秒结果写入的 sheet 名（保留 1 位小数）

FPS = 60
FRAME_MS = 1000 / FPS        # 每帧时长（毫秒），30fps 时约 33.33ms


def frame_str_to_ms(text):
    """把「秒.帧」文本转成毫秒（四舍五入整数）。

    text 必须是字符串，小数点后整段当作帧数。
    "1.10" -> 1*1000 + 10*FRAME_MS
    """
    s = str(text).strip()
    if "." not in s:
        # 没有小数点表示整秒、0 帧，例如 "13" -> 13 秒 0 帧
        sec_part, frame_part = s, "0"
    else:
        sec_part, frame_part = s.split(".", 1)
    if not (sec_part.isdigit() and frame_part.isdigit()):
        raise ValueError(f"格式错误（非纯数字）：{text!r}")
    sec = int(sec_part)
    frame = int(frame_part)   # "10" -> 10，不经过 float，帧数不丢失
    return round(sec * 1000 + frame * FRAME_MS)


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    src = wb[SOURCE_SHEET] if SOURCE_SHEET else wb.worksheets[0]

    # 重建结果 sheet（不动原始数据）
    for name in (RESULT_SHEET, RESULT_SHEET_SEC):
        if name in wb.sheetnames:
            del wb[name]
    out_ms = wb.create_sheet(RESULT_SHEET)       # 毫秒
    out_sec = wb.create_sheet(RESULT_SHEET_SEC)  # 秒，保留 1 位小数

    # 逐单元格转换，结果写到与原表完全相同的行列位置
    for row in src.iter_rows():
        for cell in row:
            value = cell.value
            if value is None or str(value).strip() == "":
                continue
            if cell.row == 1:
                # 表头原样保留
                out_ms.cell(row=cell.row, column=cell.column, value=value)
                out_sec.cell(row=cell.row, column=cell.column, value=value)
                continue
            try:
                ms = frame_str_to_ms(value)
            except ValueError as e:
                header = src.cell(row=1, column=cell.column).value
                col_name = header if header is not None else "?"
                raise ValueError(f"第「{col_name}」列 第{cell.row}行 {e}") from e
            out_ms.cell(row=cell.row, column=cell.column, value=ms)
            out_sec.cell(row=cell.row, column=cell.column, value=round(ms / 1000, 1))

    wb.save(INPUT_FILE)
    print(f"转换完成，已写入「{RESULT_SHEET}」(毫秒) 和「{RESULT_SHEET_SEC}」(秒)，位置与原表对齐。")


if __name__ == "__main__":
    main()
