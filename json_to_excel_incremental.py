"""
json_to_excel_incremental.py
用法：
    from json_to_excel_incremental import append_to_excel

    append_to_excel({
        'questions': ['送礼用的香水，有哪些品牌比较好'],
        'answers': ['https://...', 'https://...'],
        'count': 2
    })

每次调用会把数据追加到 OUTPUT_PATH 指定的 Excel 文件中。
相同问题写入同一列，不重复创建列头。
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_PATH = "youqin.xlsx"  # ← 可修改为你想要的路径

# ── 样式 ─────────────────────────────────────────────────────────────────────
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", start_color="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="DCE6F1")
DEF_FILL     = PatternFill("solid", start_color="FFFFFF")
CELL_ALIGN   = Alignment(vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _apply_header_style(cell):
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border    = BORDER


def _apply_cell_style(cell, row_idx):
    cell.font      = CELL_FONT
    cell.fill      = ALT_FILL if row_idx % 2 == 0 else DEF_FILL
    cell.alignment = CELL_ALIGN
    cell.border    = BORDER


def _refresh_row_numbers(ws, data_col_count):
    """重新写入第 A 列的序号（行数可能因追加而变化）。"""
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        cell = ws.cell(row=r, column=1, value=r - 1)
        cell.font      = CELL_FONT
        cell.fill      = ALT_FILL if r % 2 == 0 else DEF_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
        ws.row_dimensions[r].height = 20


def _init_workbook():
    """新建一个带序号列的空白工作簿。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "问答数据"
    ws.freeze_panes = "B2"

    # 序号列
    ws.column_dimensions["A"].width = 6
    cell = ws.cell(row=1, column=1, value="#")
    _apply_header_style(cell)
    ws.row_dimensions[1].height = 45

    return wb, ws


def _get_question_col_map(ws):
    """读取第 1 行，返回 {问题文本: 列号} 的映射（跳过第 1 列序号列）。"""
    mapping = {}
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            mapping[val] = col
    return mapping


def append_to_excel(data: dict, output_path: str = OUTPUT_PATH):
    """
    将一条 JSON 记录追加到 Excel。

    Parameters
    ----------
    data : dict
        格式: {'questions': [...], 'answers': [...], 'count': int}
    output_path : str
        目标 Excel 文件路径，不存在则自动创建。
    """
    questions = data.get("questions", [])
    answers   = data.get("answers", [])

    if not questions or not answers:
        print("⚠️  questions 或 answers 为空，跳过。")
        return

    # 加载或新建工作簿
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        ws = wb.active
    else:
        wb, ws = _init_workbook()

    q_col_map = _get_question_col_map(ws)

    for question in questions:
        # 确定该问题所在列（已有则复用，没有则新建）
        if question in q_col_map:
            col = q_col_map[question]
        else:
            col = ws.max_column + 1
            q_col_map[question] = col

            # 写列头
            header_cell = ws.cell(row=1, column=col, value=question)
            _apply_header_style(header_cell)
            ws.column_dimensions[
                header_cell.column_letter
            ].width = 50
            ws.row_dimensions[1].height = 45

        # 找到该列当前最后一个有数据的行
        last_row = 1
        for r in range(ws.max_row, 1, -1):
            if ws.cell(row=r, column=col).value is not None:
                last_row = r
                break

        # 追加 answers
        for answer in answers:
            next_row = last_row + 1
            cell = ws.cell(row=next_row, column=col, value=answer)
            _apply_cell_style(cell, next_row)
            last_row = next_row

    # 刷新序号列
    _refresh_row_numbers(ws, ws.max_column)

    wb.save(output_path)
    print(f"✅ 已写入 [{output_path}]  问题: {questions}  新增 {len(answers)} 条答案")


# ── 直接运行时的演示 ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    demo_records = [
        {
            "questions": ["送礼用的香水，有哪些品牌比较好"],
            "answers": [
                "https://www.doubao.com/thread/a64aa9a4940b3",
                "https://www.doubao.com/thread/aa701bcc9e1e6",
            ],
            "count": 2,
        },
        {
            "questions": ["适合送女朋友的香水有哪些"],
            "answers": [
                "https://www.doubao.com/thread/a1f5a8f184091",
                "https://www.doubao.com/thread/a0c6e8239551e",
            ],
            "count": 2,
        },
        # 相同问题再次写入 → 合并到同一列
        {
            "questions": ["送礼用的香水，有哪些品牌比较好"],
            "answers": [
                "https://www.doubao.com/thread/NEW_001",
                "https://www.doubao.com/thread/NEW_002",
            ],
            "count": 2,
        },
    ]

    TARGET = "/mnt/user-data/outputs/香水问答_增量.xlsx"
    if os.path.exists(TARGET):
        os.remove(TARGET)  # 演示时清空旧文件

    for record in demo_records:
        append_to_excel(record, output_path=TARGET)
